"""
excel_dashboard.py
Exports AML Fraud Analytics summary to a formatted Excel workbook
with multiple sheets — ready for stakeholder reporting.
"""

import pandas as pd
import sqlite3
import os

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
    from openpyxl.chart.series import DataPoint
except ImportError:
    import subprocess
    subprocess.run(["pip", "install", "openpyxl", "--break-system-packages", "-q"])
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference

conn = sqlite3.connect("data/aml_fraud.db")
txn     = pd.read_sql("SELECT * FROM fact_transactions_enriched", conn)
cust    = pd.read_sql("SELECT * FROM dim_customers", conn)
alerts  = pd.read_sql("SELECT * FROM aml_alerts", conn)
cs      = pd.read_sql("SELECT * FROM customer_summary", conn)
conn.close()

txn["date"] = pd.to_datetime(txn["date"])

OUTPUT = "dashboard/AML_Fraud_Analytics_Report.xlsx"
os.makedirs("dashboard", exist_ok=True)

with pd.ExcelWriter(OUTPUT, engine="openpyxl") as writer:

    # ── Sheet 1: Executive Summary ─────────────────────────────────────────
    summary = pd.DataFrame({
        "Metric": [
            "Total Transactions", "Total Fraud Cases", "Fraud Rate (%)",
            "Total Transaction Volume ($)", "Total Fraud Amount ($)",
            "Avg Fraud Transaction ($)", "Structuring Cases",
            "AML Alerts Generated", "Customers Flagged"
        ],
        "Value": [
            f"{len(txn):,}",
            f"{txn['is_fraud'].sum():,}",
            f"{txn['is_fraud'].mean()*100:.2f}%",
            f"${txn['amount'].sum():,.0f}",
            f"${txn[txn['is_fraud']==1]['amount'].sum():,.0f}",
            f"${txn[txn['is_fraud']==1]['amount'].mean():,.0f}",
            f"{txn['is_structuring'].sum():,}",
            f"{len(alerts):,}",
            f"{cs[cs['fraud_count']>0]['customer_id'].nunique():,}"
        ]
    })
    summary.to_excel(writer, sheet_name="Executive Summary", index=False)

    # ── Sheet 2: Fraud by Transaction Type ────────────────────────────────
    by_type = txn.groupby("transaction_type").agg(
        Total_Transactions=("is_fraud", "count"),
        Fraud_Cases=("is_fraud", "sum"),
        Fraud_Rate_Pct=("is_fraud", "mean"),
        Total_Amount=("amount", "sum"),
        Avg_Fraud_Amount=("amount", lambda x: x[txn.loc[x.index, "is_fraud"]==1].mean())
    ).reset_index()
    by_type["Fraud_Rate_Pct"] = (by_type["Fraud_Rate_Pct"] * 100).round(2)
    by_type["Total_Amount"] = by_type["Total_Amount"].round(2)
    by_type = by_type.sort_values("Fraud_Rate_Pct", ascending=False)
    by_type.to_excel(writer, sheet_name="By Transaction Type", index=False)

    # ── Sheet 3: Fraud by Country ─────────────────────────────────────────
    by_country = txn.groupby("country").agg(
        Total_Transactions=("is_fraud", "count"),
        Fraud_Cases=("is_fraud", "sum"),
        Fraud_Rate_Pct=("is_fraud", "mean"),
        Total_Amount=("amount", "sum")
    ).reset_index()
    by_country["Fraud_Rate_Pct"] = (by_country["Fraud_Rate_Pct"] * 100).round(2)
    by_country = by_country.sort_values("Fraud_Rate_Pct", ascending=False)
    by_country.to_excel(writer, sheet_name="By Country", index=False)

    # ── Sheet 4: Monthly Trend ────────────────────────────────────────────
    txn["Year_Month"] = txn["date"].dt.to_period("M").astype(str)
    monthly = txn.groupby("Year_Month").agg(
        Total_Transactions=("is_fraud", "count"),
        Fraud_Cases=("is_fraud", "sum"),
        Fraud_Rate_Pct=("is_fraud", "mean"),
        Total_Amount=("amount", "sum"),
        Fraud_Amount=("amount", lambda x: x[txn.loc[x.index, "is_fraud"]==1].sum())
    ).reset_index()
    monthly["Fraud_Rate_Pct"] = (monthly["Fraud_Rate_Pct"] * 100).round(2)
    monthly.to_excel(writer, sheet_name="Monthly Trend", index=False)

    # ── Sheet 5: AML Alerts ───────────────────────────────────────────────
    alerts_out = alerts[["transaction_id","customer_id","date","amount",
                          "transaction_type","country","fraud_score","alert_reason"]].copy()
    alerts_out["amount"] = alerts_out["amount"].round(2)
    alerts_out["fraud_score"] = alerts_out["fraud_score"].round(2)
    alerts_out.to_excel(writer, sheet_name="AML Alerts", index=False)

    # ── Sheet 6: Customer Risk Summary ────────────────────────────────────
    cust_risk = cs.merge(cust[["customer_id","risk_tier","kyc_status","is_pep","country"]], on="customer_id")
    cust_risk["fraud_rate"] = (cust_risk["fraud_rate"] * 100).round(2)
    cust_risk["total_amount"] = cust_risk["total_amount"].round(2)
    cust_risk = cust_risk.sort_values("fraud_count", ascending=False)
    cust_risk.to_excel(writer, sheet_name="Customer Risk", index=False)

# ── Apply formatting ───────────────────────────────────────────────────────────
wb = load_workbook(OUTPUT)

HEADER_FILL = PatternFill("solid", fgColor="1B3A6B")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
ALT_FILL    = PatternFill("solid", fgColor="EAF0FB")
BORDER_SIDE = Side(style="thin", color="CCCCCC")
CELL_BORDER = Border(left=BORDER_SIDE, right=BORDER_SIDE, top=BORDER_SIDE, bottom=BORDER_SIDE)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for cell in ws[1]:
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        for cell in row:
            cell.fill   = fill
            cell.border = CELL_BORDER
            cell.alignment = Alignment(horizontal="center")
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)
    ws.freeze_panes = "A2"

wb.save(OUTPUT)
print(f"✅ Excel report saved: {OUTPUT}")
